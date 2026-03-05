#!/usr/bin/env python3
"""Generate Site Daily Runsheet Excel template."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.environ.get("OUTPUT_DIR", ".") + "/site-daily-runsheet.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Arial", bold=True, color="2F5496", size=12)
LABEL_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
INPUT_FILL = PatternFill(start_color="FFFFEB", end_color="FFFFEB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def style_cell(ws, row, col, value="", font=BODY_FONT, fill=None, merge_end=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)
        for c in range(col, merge_end + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
    return cell


def header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = style_cell(ws, row, col_start + i, h, HEADER_FONT, HEADER_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")


def input_rows(ws, start_row, num_rows, num_cols, col_start=1):
    for r in range(start_row, start_row + num_rows):
        for c in range(col_start, col_start + num_cols):
            style_cell(ws, r, c, "", BODY_FONT, INPUT_FILL)


def create_runsheet(wb):
    ws = wb.active
    ws.title = "Daily Runsheet"
    ws.sheet_properties.tabColor = "2F5496"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Column widths (A-H for landscape A4)
    widths = [20, 20, 18, 16, 14, 14, 14, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    row = 1
    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "DAILY SITE RUNSHEET"
    t.font = Font(name="Arial", bold=True, size=18, color="2F5496")
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Project info section
    row = 3
    labels = [
        ("Project Name:", 1, 2, 3), ("Site Address:", 5, 6, 8),
        ("Date:", 1, 2, 3), ("Weather:", 5, 6, 8),
        ("Site Manager:", 1, 2, 3), ("Shift:", 5, 6, 8),
    ]
    for i in range(0, len(labels), 2):
        l1 = labels[i]
        l2 = labels[i + 1]
        style_cell(ws, row, l1[1], l1[0], LABEL_FONT)
        style_cell(ws, row, l1[2], "", BODY_FONT, INPUT_FILL, l1[3])
        style_cell(ws, row, l2[1], l2[0], LABEL_FONT)
        style_cell(ws, row, l2[2], "", BODY_FONT, INPUT_FILL, l2[3])
        row += 1

    # Section 1: Personnel On-Site
    row += 1
    style_cell(ws, row, 1, "SECTION 1: PERSONNEL ON-SITE", SECTION_FONT, merge_end=8)
    row += 1
    header_row(ws, row, ["Name", "Company", "Role", "Trade", "Start", "Finish", "Hours", "Notes"])
    row += 1
    input_rows(ws, row, 12, 8)
    row += 12

    # Personnel total
    style_cell(ws, row, 1, "Total Personnel:", LABEL_FONT)
    style_cell(ws, row, 2, "", BODY_FONT, INPUT_FILL)
    style_cell(ws, row, 5, "Total Hours:", LABEL_FONT)
    style_cell(ws, row, 6, f"=SUM(G{row - 12}:G{row - 1})", BODY_FONT, INPUT_FILL)

    # Section 2: Plant & Equipment
    row += 2
    style_cell(ws, row, 1, "SECTION 2: PLANT & EQUIPMENT", SECTION_FONT, merge_end=8)
    row += 1
    header_row(ws, row, ["Item/Description", "Supplier", "Docket No.", "Hours Used", "Status", "Operator", "Location", "Notes"])
    row += 1
    input_rows(ws, row, 8, 8)
    row += 8

    # Section 3: Work Completed Today
    row += 1
    style_cell(ws, row, 1, "SECTION 3: WORK COMPLETED TODAY", SECTION_FONT, merge_end=8)
    row += 1
    header_row(ws, row, ["Area / Zone", "Description of Work", "", "Trade", "% Complete", "Planned %", "Variance", "Notes"])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 1
    for r in range(row, row + 8):
        style_cell(ws, r, 1, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 2, "", BODY_FONT, INPUT_FILL, 3)
        for c in [4, 5, 6, 8]:
            style_cell(ws, r, c, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 7, f"=E{r}-F{r}", BODY_FONT)
    row += 8

    # Section 4: Safety Observations
    row += 1
    style_cell(ws, row, 1, "SECTION 4: SAFETY OBSERVATIONS", SECTION_FONT, merge_end=8)
    row += 1
    header_row(ws, row, ["Time", "Observation", "", "Risk Level", "Action Required", "", "Responsible", "Closed?"])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 1
    for r in range(row, row + 6):
        style_cell(ws, r, 1, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 2, "", BODY_FONT, INPUT_FILL, 3)
        style_cell(ws, r, 4, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 5, "", BODY_FONT, INPUT_FILL, 6)
        style_cell(ws, r, 7, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 8, "", BODY_FONT, INPUT_FILL)
    row += 6

    # Section 5: Delays/Issues
    row += 1
    style_cell(ws, row, 1, "SECTION 5: DELAYS / ISSUES", SECTION_FONT, merge_end=8)
    row += 1
    header_row(ws, row, ["Description", "", "Impact (hrs)", "Cause", "Action Taken", "", "Responsible", "Resolved?"])
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 1
    for r in range(row, row + 5):
        style_cell(ws, r, 1, "", BODY_FONT, INPUT_FILL, 2)
        style_cell(ws, r, 3, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 4, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 5, "", BODY_FONT, INPUT_FILL, 6)
        style_cell(ws, r, 7, "", BODY_FONT, INPUT_FILL)
        style_cell(ws, r, 8, "", BODY_FONT, INPUT_FILL)
    row += 5

    # Section 6: Materials Delivered
    row += 1
    style_cell(ws, row, 1, "SECTION 6: MATERIALS DELIVERED", SECTION_FONT, merge_end=8)
    row += 1
    header_row(ws, row, ["Item", "Quantity", "Unit", "Supplier", "Docket No.", "Received By", "Storage Location", "Condition"])
    row += 1
    input_rows(ws, row, 6, 8)
    row += 6

    # Footer: Sign-off
    row += 2
    style_cell(ws, row, 1, "Prepared By:", LABEL_FONT)
    style_cell(ws, row, 2, "", BODY_FONT, INPUT_FILL, 3)
    style_cell(ws, row, 5, "Date:", LABEL_FONT)
    style_cell(ws, row, 6, "", BODY_FONT, INPUT_FILL)
    row += 1
    style_cell(ws, row, 1, "Signature:", LABEL_FONT)
    style_cell(ws, row, 2, "", BODY_FONT, INPUT_FILL, 3)
    style_cell(ws, row, 5, "Reviewed By:", LABEL_FONT)
    style_cell(ws, row, 6, "", BODY_FONT, INPUT_FILL, 8)

    # Print settings
    ws.print_area = f"A1:H{row}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def create_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = "70AD47"

    instructions = [
        ("SITE DAILY RUNSHEET - INSTRUCTIONS", True, 16),
        ("", False, 10),
        ("Purpose", True, 13),
        ("This template captures daily site activities for construction projects.", False, 10),
        ("Complete it at the end of each day and submit to the Project Manager.", False, 10),
        ("", False, 10),
        ("Section Guide", True, 13),
        ("", False, 10),
        ("Section 1: Personnel On-Site", True, 11),
        ("Record all personnel present on site including subcontractors and visitors.", False, 10),
        ("Start/Finish times should be in 24-hour format (e.g., 07:00, 15:30).", False, 10),
        ("Hours column should show total hours worked (excluding breaks).", False, 10),
        ("", False, 10),
        ("Section 2: Plant & Equipment", True, 11),
        ("List all plant and equipment used on site that day.", False, 10),
        ("Status options: In Use, Idle, Breakdown, Maintenance.", False, 10),
        ("Record docket numbers for hired equipment.", False, 10),
        ("", False, 10),
        ("Section 3: Work Completed Today", True, 11),
        ("Describe work by area/zone for easy tracking against the programme.", False, 10),
        ("% Complete should reflect cumulative progress, not just today's work.", False, 10),
        ("Variance auto-calculates (Actual % minus Planned %).", False, 10),
        ("", False, 10),
        ("Section 4: Safety Observations", True, 11),
        ("Record all safety observations, near-misses, and positive observations.", False, 10),
        ("Risk Level: Low, Medium, High, Critical.", False, 10),
        ("All observations must have an action and responsible person.", False, 10),
        ("", False, 10),
        ("Section 5: Delays / Issues", True, 11),
        ("Document any delays or issues affecting the programme.", False, 10),
        ("Quantify the impact in hours where possible.", False, 10),
        ("Include weather delays, material shortages, coordination issues.", False, 10),
        ("", False, 10),
        ("Section 6: Materials Delivered", True, 11),
        ("Record all material deliveries with docket numbers.", False, 10),
        ("Note condition on arrival (Good, Damaged, Incomplete).", False, 10),
        ("", False, 10),
        ("Tips for Daily Use", True, 13),
        ("- Fill in the header info (project, date, weather) first thing each morning", False, 10),
        ("- Update personnel as they sign in/out throughout the day", False, 10),
        ("- Note safety observations as they happen, not at end of day", False, 10),
        ("- Take photos to support any observations or issues noted", False, 10),
        ("- Submit by 6:00 PM each day to the Project Manager", False, 10),
        ("- Keep a copy for site records", False, 10),
    ]

    for i, (text, bold, size) in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", bold=bold, size=size,
                         color="2F5496" if bold and size > 11 else "000000")

    ws.column_dimensions["A"].width = 80


def main():
    wb = openpyxl.Workbook()
    create_runsheet(wb)
    create_instructions(wb)
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
